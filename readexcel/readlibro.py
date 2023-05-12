import openpyxl import load_workbook

# Leer un archivo existente
workbook = openpyxl.load_workbook('../datos_usuarios.xlsx')

# obtenemos la pestaña/hoja activa (nada mas abrir es la primera)
hoja = workbook.active
ultimafilahoja = hoja.get_highest_row()
ultimafila = 1
# mostramos la celda B1 = 1:2
#print(hoja.cell(row=1, column=2).value)
name=hoja.cell(row=1, column=2).value
#print(hoja['B1'].value)
#Recorrer la hoja
i=2 #Empezamos en 2 por las cabeceras
while i <= hoja.max_row:
    print(i)
    i += 1

# Crear un nuevo archivo
#new_workbook = openpyxl.Workbook()
#new_workbook.save('../datos_usuarios__.xlsx')

#crear el archivo plano
with open("../query.txt","w") as file:
    file.write(name+"\n")